import tkinter as tk
from tkinter import PhotoImage, filedialog, messagebox
import pandas as pd
import json
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
import sys

def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def change_file_extension(filename, new_ext):
    return f"{'.'.join(filename.split('.')[:-1])}{new_ext}"

def upload_file():
    file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
    if file_path:
        input_file.set(file_path)
        if not output_file.get():
            new_ext = '.xlsx' if file_type.get() == 'Excel' else '.csv'
            output_file.set(change_file_extension(file_path, new_ext))

def save_file_as():
    ext = '.xlsx' if file_type.get() == 'Excel' else '.csv'
    file_path = filedialog.asksaveasfilename(defaultextension=ext, filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
    if file_path:
        output_file.set(file_path)
        file_type.set('Excel' if file_path.endswith('.xlsx') else 'CSV')

def convert_json():
    try:
        with open(input_file.get(), 'r') as file:
            data = json.load(file)
        
        output_path = output_file.get()
        
        if is_questionnaire.get():
            sections = data['Sections']
            metadata = {key: value for key, value in data.items() if not isinstance(value, list) and key != 'Sections'}
            df = pd.DataFrame()
            data_rows = []
            for section in sections:
                section_name = section['NameSection']
                for question in section['Questions']:
                    question_text = question['QuestionText']
                    for response in question['RespondentAnswers']:
                        for answer in response['LikertLearningStudentAnswers']:
                            row_data = {
                                'Section': section_name,
                                'Question': question_text,
                                'Teacher ID': answer['IdUserTeacher'],
                                'Perception': answer.get('Perception', ''),
                                'Value': answer.get('Value', '')
                            }
                            row_data.update(metadata)
                            data_rows.append(row_data)
            
            df = pd.DataFrame(data_rows)
        else:
            df = pd.json_normalize(data)

        if file_type.get() == 'Excel':
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
                worksheet = writer.sheets['Sheet1']

                if bold_header.get():
                    bold_font = Font(bold=True)
                    for col_num, value in enumerate(df.columns.values):
                        col_letter = get_column_letter(col_num + 1)
                        worksheet[f'{col_letter}1'].font = bold_font

                if auto_fit.get():
                    for col_num, col in enumerate(df.columns.values):
                        max_length = max(df[col].astype(str).map(len).max(), len(col))
                        worksheet.column_dimensions[get_column_letter(col_num + 1)].width = max_length + 2
        else:
            # Save as CSV
            df.to_csv(output_path, index=False)

        # Show success message
        messagebox.showinfo("Success", f"File saved successfully at {output_path}")
    except Exception as e:
        # Show error message
        messagebox.showerror("Error", str(e))

def update_output_file(*args):
    if input_file.get() and output_file.get():
        new_ext = '.xlsx' if file_type.get() == 'Excel' else '.csv'
        output_file.set(change_file_extension(output_file.get(), new_ext))

# Main window
root = tk.Tk()
root.title("BEBEK CONVERTER")
iconPath = resource_path('duck.ico')
root.iconbitmap(iconPath)
root.geometry("400x260")
root.resizable(False, False)

input_file = tk.StringVar()
output_file = tk.StringVar()
bold_header = tk.BooleanVar()
auto_fit = tk.BooleanVar()
is_questionnaire = tk.BooleanVar()
file_type = tk.StringVar(value="Excel")
file_type.trace_add("write", update_output_file)

frame = tk.Frame(root, padx=10, pady=10)
frame.pack(fill=tk.BOTH, expand=True)

tk.Label(frame, text="JSON File:").grid(row=0, column=0, sticky=tk.W, pady=5)
tk.Entry(frame, textvariable=input_file, width=40).grid(row=0, column=1, pady=5)
tk.Button(frame, text="Browse", command=upload_file, width=10).grid(row=0, column=2, padx=5, pady=5)

tk.Label(frame, text="Save As:").grid(row=1, column=0, sticky=tk.W, pady=5)
tk.Entry(frame, textvariable=output_file, width=40).grid(row=1, column=1, pady=5)
tk.Button(frame, text="Save As", command=save_file_as, width=10).grid(row=1, column=2, padx=5, pady=5)

tk.Label(frame, text="File Type:").grid(row=2, column=0, sticky=tk.W, pady=5)
tk.Radiobutton(frame, text="Excel", variable=file_type, value="Excel").grid(row=2, column=1, sticky=tk.W)
tk.Radiobutton(frame, text="CSV", variable=file_type, value="CSV").grid(row=2, column=1)

tk.Checkbutton(frame, text="Bold Headers", variable=bold_header).grid(row=3, column=1, sticky=tk.W, pady=5)
tk.Checkbutton(frame, text="Auto Fit Columns", variable=auto_fit).grid(row=4, column=1, sticky=tk.W, pady=5)
tk.Checkbutton(frame, text="Is Questionnaire", variable=is_questionnaire).grid(row=5, column=1, sticky=tk.W, pady=5)

tk.Button(frame, text="Convert", command=convert_json, width=20).grid(row=6, column=1, pady=10)

root.mainloop()
